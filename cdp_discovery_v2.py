import openpyxl
import os
import re
import sys
import time
import netmiko
import paramiko
from netmiko import ConnectHandler

excelFile = "TestDevice.xlsx"
timeStamp = time.strftime("__[%Y.%b.%d].[%I.%M.%S.%p]")

device_type = "cisco_ios"
username = input('Enter the Username: ')
password = input('Enter the Password: ')
ip = input('Enter the seed device IP: ')

match_set = {""}

def open_file (excelFile):
    wb = openpyxl.load_workbook(excelFile)
    ws = wb.worksheets [0]
    devices = []
    row = 2
    column = 1
    while row <= 3:
        device_name_xl = ws.cell(row, column).value
        device_para_xl = {'device_type': ws.cell(row, column+1).value,'ip':ws.cell(row, column+2).value,'username': ws.cell(row, column+3).value,'password': ws.cell(row, column+4).value}
        devices.append(device_para_xl)
        row+= 1
    return devices

def connect (devices):
    for device_para in devices:
        try:
            net_connect = ConnectHandler(**device_para)
            Device_Name = net_connect.find_prompt()
            # Device_Name = Device_Name.replace('#', '')
            output = net_connect.send_command("show cdp neighbors detail | in Device ID:.+\\.com|IP address:|Platform|Version")
            find_matches(Device_Name, output)
            net_connect.disconnect()
        except paramiko.ssh_exception.AuthenticationException:
            print("\n>>>> Device IP: {0} <<<<".format(device_para.get('ip')))
            print('I\'ve got an Authentication issue')
            print("\t>>>>>>>>> End <<<<<<<<<\n")
            pass
        except netmiko.ssh_exception.NetMikoTimeoutException:
            print("\n>>>> Device IP: {0} <<<<".format(device_para.get('ip')))
            print('This Device is UnReachable')
            print("\t>>>>>>>>> End <<<<<<<<<\n")
            pass
        except Exception as e:
            print("\n>>>> Device IP: {0} <<<<".format(device_para.get('ip')))
            print('Maybe it is an Invalid command for me')
            print("\t>>>>>>>>> End <<<<<<<<<\n"+ str(e))
            pass

def find_matches (Device_Name, output):
    regex = '''(Device ID: .+\s+IP address:\s+\d+\.\d+\.\d+\.\d+\s*Platform:.+\s*Capabilities:.+\s*Version :\s*.+Version \S+)'''
    pattern = re.compile(regex)
    matches = pattern.findall(output)
    matches = set(matches)
    if len(matches) < 1:
        print(">>>> Device Name: {} <<<<".format(Device_Name))
        print("I'm sorry, I can't list my Neighbors")
    else:
        for match in matches:
            match = re.sub('\.\w+\.com', '', match)
            match = re.sub('(Platform:\s*.+),\s*(Capabilities:\s*.+)', '  \\1\n  \\2', match)
            match = re.sub('(Version :)\n(.+),', '  \\1 \\2', match)
            if match not in match_set:
                match_set.add(match)
                # print(match)
                match_lines_list = str.splitlines(match)
                device_ip = re.findall(re.compile('\d+\.\d+\.\d+\.\d+'), match_lines_list[1])
                device_para = [
                    {'device_type': device_type, 'ip': device_ip.pop(), 'username': username, 'password': password}]
                connect(device_para)

if __name__ == '__main__':
    # devices_xl = open_file(excelFile)
    device_para = [{'device_type': device_type, 'ip': ip, 'username': username, 'password': password}]
    sys.stdout = open('matches_set@' + str(timeStamp) + ".txt", 'w')
    connect(device_para)
    for m in sorted(match_set):
        print('\n'+m)
    sys.stdout.close()